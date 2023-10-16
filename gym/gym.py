import pandas as pd
import numpy as np
import itertools
import math
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from os import path

def import_data(return_dicts=False):
    """ Imports and combines data for both days of competition, calculates average by athlete, 
    and returns DataFrame(s) of all data (optionally second vault data)."""
    day1 = pd.read_excel(r'./gym/data/2023 US Championships Results.xlsx', sheet_name='Prelims')
    day2 = pd.read_excel(r'./gym/data/2023 US Championships Results.xlsx', sheet_name='Finals')
    # merge day 1 and day 2 results
    AA = day1.merge(day2, how='inner', on='Name', suffixes=['_day1','_day2'])
    # calculate averages for each athlete and event (including AA)
    for event in ['AA', 'Vault', 'Bars', 'Beam', 'Floor']:
        AA[f'{event}_avg'] = AA[[f'{event}_day1', f'{event}_day2']].mean(axis=1)
    # sort DF by descending all around average
    AA.sort_values(by='AA_avg', ascending=False, inplace=True)

    # optionally return dictionaries mapping athletes to a color and an int
    if return_dicts:
        # create dict mapping each athlete to a unique color
        name_color = AA[['Name', 'Color']].set_index('Name')['Color'].to_dict() 
        # create dict to encode athlete names as ints to speed up later processes
        name_int = AA.reset_index().set_index('Name')['index'].to_dict()
        return AA, name_color, name_int

    return AA
   

def top_team_scores(AA, occ):
    """ For each possible 5-member team, calculates the team score using the top 3 scores
    on each event among the team members, simulating a 3-up, 3-count competition.
    Returns dataframe with Team ID, members, and score, sorted by team score;
    matrix of top 3 scores on each event for all possible teams, 
    nested list of athletes who received the aforementioned scores. """
    num_athletes = len(AA)
    # create structures to store each id's members, counting scores and athletes for each event, total team score
    team_id = pd.DataFrame(columns=['Team ID', 'Team Score', 'Member 1', 'Member 2', 'Member 3', 'Member 4', 'Member 5'])
    #A- team_scores = np.zeros(math.comb(num_athletes, 5))
    counting_scores = np.zeros((math.comb(num_athletes, 5), 4, 3))
    counting_names = [ [ [ [] for  i in range(3) ] for i in range(4)] for i in range(math.comb(num_athletes, 5))]
    # iterate through all 5-member team combinations
    for i, row in enumerate(itertools.combinations(AA.values, 5)):
        team = pd.DataFrame(row, columns=['Name', 'Vault_day1', 'Bars_day1', 'Beam_day1', 'Floor_day1', 'AA_day1',
        'Color', 'Vault_day2', 'Bars_day2', 'Beam_day2', 'Floor_day2', 'AA_day2',
        'AA_avg', 'Vault_avg', 'Bars_avg', 'Beam_avg', 'Floor_avg'])  
        score = 0 # team score
        # iterate over all events
        for ii, event in enumerate(['Vault', 'Bars', 'Beam', 'Floor']):
            # specify event and occasion (day or average)
            event_occ = f'{event}_{occ}'
            # create temporary df of the team sorted by event score
            sort = team[[f'Vault_{occ}', f'Bars_{occ}', f'Beam_{occ}', f'Floor_{occ}', 'Name']].sort_values(by=event_occ, ascending=False)[:3]
            # store top 3 scores and associated athlete names
            counting_scores[i,ii,:] = flatten([[sort[event_occ].iloc[j]] for j in range(3)])
            counting_names[i][ii] = [[sort['Name'].iloc[j]] for j in range(3)]
            score += counting_scores[i,ii,:].sum()

        team_members = [i] + [round(score,2)] + team['Name'].tolist()
        team_id.loc[i] = team_members

    return team_id, counting_scores, counting_names
    

def write_team_scores_to_excel(counting_names, counting_scores, sheet_name, num_athletes=28):
    """ Writes counting routine data for each possible 5-member team to Excel to avoid re-running all combinations.
     Sheet name specifies occasion (day or average). Called by run_team_combinations. """
    # construct dataframe of all counting scores for each team
    team_data = {'Team ID': np.repeat(np.array([i for i in range(math.comb(num_athletes, 5))]), 12), # 12 counting scores
            'Event': np.tile(np.repeat(np.array(['Vault', 'Bars', 'Beam', 'Floor']), 3), math.comb(num_athletes, 5)),
            'Score_Rank': np.tile(np.array([1,2,3]), 4*math.comb(num_athletes, 5)),
            'Name': np.reshape(np.array(counting_names), -1),
            'Score': np.reshape(counting_scores, -1)}

    # create blank excel file if doesn't already exist       
    if path.exists(r'./gym/data/Highest Scoring Teams.xlsx') == False:
        wb = Workbook()
        wb.save(r'./gym/data/Highest Scoring Teams.xlsx')

    # write dataframe to sheet        
    wb = load_workbook(r'./gym/data/Highest Scoring Teams.xlsx')
    sheet = wb.create_sheet(f'{sheet_name}')
    sheet.append(['Team ID', 'Event', 'Score_Rank', 'Name', 'Score'])
    for r in dataframe_to_rows(pd.DataFrame(team_data), index=False, header=False):
            sheet.append(r)
    
    wb.save(r'./gym/data/Highest Scoring Teams.xlsx')

    return pd.DataFrame(team_data)

def run_team_combinations(AA):
    """ Runs all possible team combinations with day 1, day 2, and average scores.
    Writes counting routine data to Excel, Team ID, score, and members to CSV. """
    sheet_names = ['Day 1', 'Day 2', 'Average']
    
    # loop over all occasions
    for i, occ in enumerate(['day1', 'day2', 'avg']):
        # calculate scores for all possible team combinations
        team_id, counting_scores, counting_names = top_team_scores(AA, occ)
        # write team member data to CSV
        team_id.to_csv(f'./gym/data/teams/{sheet_names[i]} Teams.csv', index=False)
        # write counting score data to Excel
        write_team_scores_to_excel(counting_names, counting_scores, sheet_names[i])

def import_counting_scores(sheet_name):
    """ Used to import counting scores for the occasion specified by sheet name which are returned as a DataFrame.
    Checks for teams with all of the same counting scores, removes them from the aforementioned DataFrame, and returns all team data along
    with Team IDs of the removed equivalent teams. """
    # import counting score data 
    scores = pd.read_excel(r'./gym/data/Highest Scoring Teams.xlsx', sheet_name=sheet_name)
    # import team member combinations and sort by team score
    teams = pd.read_csv(f'./gym/data/teams/{sheet_name} Teams.csv')
    teams.sort_values(by='Team Score', ascending=False, inplace=True)
    teams.reset_index(drop=True, inplace=True)
    # find and remove any teams with the same 12 counting routines
    duplicates = find_same_3up(teams, scores)
    scores, removed_teams = remove_duplicate_3up(scores, duplicates)

    return scores, teams, removed_teams

def find_same_3up(team_members, counting_scores):
    """ Iterates through all teams which share a team score to find teams that have different team members,
    but utilize the same set of 12 counting routines. Returns a lested list of arrays with Team IDs
    of teams which share duplicate routines. """
    # narrow search to only observations which have a duplicate team score
    duplicate_score_rows = team_members[team_members.duplicated(subset='Team Score', keep=False)].copy()
    # find each unique score to iterate through
    unique_scores = duplicate_score_rows['Team Score'].unique() # get all scores that multiple teams achieved
    redundant_teams = []

    # iterate across all duplicated scores
    for score in unique_scores:
        # find all team id's that had team score of score
        team_ids = duplicate_score_rows[duplicate_score_rows['Team Score'] == score]['Team ID'].values
        same_score_teams = counting_scores[counting_scores['Team ID'].isin(team_ids)].copy()
        same_score_teams.set_index('Team ID', inplace=True)
        # create dictionaries for encoding
        _, _, name_int = import_data(return_dicts=True)
        event_int = {'Vault': 0, 'Bars': 1, 'Beam': 2, 'Floor': 3}
        # encode names and events for quick comparison of 3d arrays
        same_score_teams['Athlete ID'] = same_score_teams['Name'].apply(lambda x: name_int[x])
        same_score_teams['Event ID'] = same_score_teams['Event'].apply(lambda x: event_int[x])
        counting = np.array(same_score_teams[['Athlete ID', 'Score', 'Score_Rank', 'Event ID']]).reshape((len(team_ids),12,4))
        # find duplicate teams (have same 12 counting routines)
        _, inverse, count = np.unique(counting, return_inverse=True, return_counts=True, axis=0)
        duplicate_idx = np.where(count[inverse]>1)[0]
        duplicate_teams = team_ids[duplicate_idx]
        redundant_teams.append(duplicate_teams)        

    return [ele for ele in redundant_teams if ele != []]

# remove first instance to keep one 
def remove_duplicate_3up(counting_scores, duplicates):
    """ Removes all but one observation of identified duplicates from DataFrame of all counting scores.
    Returns altered DataFrame along with a dictionary mapping the non-removed team to all its equivalent teams by ID. """
    team_ids_to_remove = duplicates.copy()
    # create dict to map each team not removed to its duplicates to be removed
    duplicates_dict = {group[0]: group[1:] for group in team_ids_to_remove}
    duplicates_to_remove = flatten(list(duplicates_dict.values()))
    # remove all but one duplicate per identical group
    counting_scores = counting_scores.drop(counting_scores[counting_scores['Team ID'].isin(duplicates_to_remove)].index)

    return counting_scores, duplicates_dict

def get_duplicates_for_top_team_table(top_team_ids, removed_teams, team_df):
    """ Scans top team ids to check if any teams had duplicates removed for before plotting.
    Returns arrays of any duplicate teams present (by ID), the team members which are constant
    among the duplicate teams, and those which are variable to annotate plots. """
    full_team_ids = []
    # search top team id's for teams which had removed duplicates to pull info for table
    for team in top_team_ids:
        if team in removed_teams:
            full_team_ids.append([[team], list(removed_teams[team])])
        else:
            full_team_ids.append([team])
    duplicate_team_ids = []
    team_constants = []
    team_variables = []
    for i, teams in enumerate(full_team_ids):
        # if a top team had duplicates, return team id, its constant members, and its variable members
        if len(teams) > 1:
            # construct dataframe of all equivalent team combinations
            team_options = team_df[team_df['Team ID'] == full_team_ids[i][0][0]].drop(['Team ID', 'Team Score'], axis=1).copy()
            team_options = pd.concat([team_options, team_df[team_df['Team ID'].isin(full_team_ids[i][1])].drop(['Team ID', 'Team Score'], axis=1)])
            # find the 4 recurring athletes and the swappable ones
            value_counts = pd.Series(flatten(team_options.values)).value_counts().to_frame()
            duplicate_team_ids.append(teams[0])
            team_constants.append(list(value_counts.iloc[:4].index))
            team_variables.append(list(value_counts.iloc[4:].index))

    if duplicate_team_ids:
        return duplicate_team_ids, team_constants, team_variables
    else:
        return 0, 0, 0

def flatten(l):
    """ Flatten nested list. """
    return [item for sublist in l for item in sublist]
